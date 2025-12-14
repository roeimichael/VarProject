import numpy as np
import pandas as pd
import yfinance as yf
from pandas_datareader import data as pdr
import datetime as dt
import openpyxl
from openpyxl.styles import PatternFill
import logging
import Var
import config

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


def check_tickers():
    """Check which tickers can be successfully fetched from Yahoo Finance."""
    tickers = get_list('alltickers')
    bad_stocks = []
    for index, ticker in enumerate(tickers):
        try:
            logger.info(f"Checking ticker {index}: {ticker}")
            data = pdr.get_data_yahoo(ticker)
        except Exception as e:
            logger.warning(f"Failed to fetch data for {ticker}: {e}")
            bad_stocks.append(ticker)
    logger.info(f"Bad stocks: {bad_stocks}")
    return bad_stocks


def get_list(filename):
    """Load a list of tickers from a text file."""
    ticker_list = []
    try:
        with open(f'./data/{filename}.txt', 'r') as fp:
            for line in fp:
                ticker = line.strip()
                if ticker:  # Skip empty lines
                    ticker_list.append(ticker)
        return ticker_list
    except FileNotFoundError:
        logger.error(f"File not found: ./data/{filename}.txt")
        raise
    except Exception as e:
        logger.error(f"Error reading file {filename}.txt: {e}")
        raise


def transform_df(df_path, net_liquidity):
    """Transform raw portfolio CSV into standardized format with sector information."""
    try:
        logger.info(f"Loading portfolio from {df_path}")
        df = pd.read_csv(df_path)
        df = df.loc[:, ~df.columns.str.contains('^Unnamed')]
        df = df.drop(['Daily P&L', 'Last', 'Change', 'Unrealized P&L', 'Market Value'], axis=1)
        df.rename(columns={'Financial Instrument': 'Symbol'}, inplace=True)
        df['Sector'] = ''
        df['Amount'] = 0.0

        for index, row in df.iterrows():
            df.at[index, 'Position'] = str(df.at[index, 'Position']).replace(',', '')
            ticker = df.at[index, 'Symbol']
            try:
                tick = yf.Ticker(ticker)
                if 'sector' in tick.info:
                    df.at[index, 'Sector'] = tick.info['sector']
                else:
                    df.at[index, 'Sector'] = 'ETF'
            except Exception as e:
                logger.warning(f"Could not fetch sector for {ticker}: {e}")
                df.at[index, 'Sector'] = 'Unknown'

        df["Position"] = pd.to_numeric(df["Position"])
        df['Amount'] = df['Position'] * df['Avg Price']
        df['Type'] = np.where(df['Amount'] > 0, 'LONG', 'SHORT')
        df = df.drop(['Avg Price', 'Position'], axis=1)
        df['Percentage'] = abs(df['Amount']) / net_liquidity
        cols = ['Symbol', 'Type', 'Amount', 'Percentage', 'Sector']
        df = df[cols]
        df.rename(columns={'Type': 'Position', 'Percentage': 'Protfilio Precentage'}, inplace=True)
        df.to_excel(config.PORTFOLIO_PATH_TRANSFORMED, index=False)
        logger.info(f"Portfolio transformed and saved to {config.PORTFOLIO_PATH_TRANSFORMED}")

    except FileNotFoundError:
        logger.error(f"Portfolio file not found: {df_path}")
        raise
    except KeyError as e:
        logger.error(f"Expected column not found in CSV: {e}")
        raise
    except Exception as e:
        logger.error(f"Error transforming portfolio: {e}")
        raise


def coloring_portfolio(filename):
    """Add color-coded VaR quality ratings to portfolio Excel file."""
    try:
        wb = openpyxl.load_workbook(filename=filename)
        ws = wb['Sheet1']
        df = pd.read_excel(filename)
        length = df.shape[0] + 2

        # Define color fills for quality ratings
        fill_cell_bad = PatternFill(patternType='solid', fgColor='FC2C03')  # Red
        fill_cell_mid = PatternFill(patternType='solid', fgColor='FFFF00')  # Yellow
        fill_cell_good = PatternFill(patternType='solid', fgColor='35FC03')  # Green

        # Determine columns based on file type
        if filename == config.ALL_TICKERS_PATH:
            var_col, qual_col = 'B', 'C'
        else:
            var_col, qual_col = 'F', 'G'

        # Color-code rows based on VaR percentiles
        # Top 33% = GOOD (green), Middle 57% = MID (yellow), Bottom 10% = BAD (red)
        for i in range(2, round(length / 3)):
            ws[f'{var_col}{i}'].fill = fill_cell_good
            ws[f'{qual_col}{i}'] = 'GOOD'
        for i in range(round(length / 3), round(9 * length / 10)):
            ws[f'{var_col}{i}'].fill = fill_cell_mid
            ws[f'{qual_col}{i}'] = 'MID'
        for i in range(round(9 * length / 10), length):
            ws[f'{var_col}{i}'].fill = fill_cell_bad
            ws[f'{qual_col}{i}'] = 'BAD'

        wb.save(filename=filename)
        logger.info(f"Color-coded quality ratings added to {filename}")

    except Exception as e:
        logger.error(f"Error coloring portfolio {filename}: {e}")
        raise


def get_var(portfolio_path, alltickers_path):
    """Calculate VaR for portfolio tickers and add quality ratings."""
    try:
        logger.info("Starting VaR calculation")
        coloring_portfolio(alltickers_path)
        alltickers_df = pd.read_excel(alltickers_path)
        alltickers_df = alltickers_df.loc[:, ~alltickers_df.columns.str.contains('^Unnamed')]
        alltickers_list = alltickers_df['Symbol'].tolist()
        portfolio_df = pd.read_excel(portfolio_path)
        portfolio_df['Var'] = 0
        portfolio_df['Qual'] = 0
        portfolio_tickers = portfolio_df['Symbol'].tolist()
        alltickers_df.set_index("Symbol", drop=False, inplace=True)
        portfolio_df.set_index("Symbol", drop=False, inplace=True)

        for tick in portfolio_tickers:
            logger.info(f"Processing VaR for {tick}")
            if tick in alltickers_list:
                # Use cached VaR value
                portfolio_df.at[tick, 'Var'] = alltickers_df.at[tick, 'Var']
                portfolio_df.at[tick, 'Qual'] = alltickers_df.at[tick, 'Qual']
            else:
                # Calculate new VaR
                try:
                    data = pdr.get_data_yahoo([tick], start=config.START_DATE_FOR_VAR, end=dt.date.today())['Close']
                    cur_var = Var.calc_var(config.INITIAL_INVESTMENT, np.array(config.WEIGHTS), data)
                    alltickers_df.loc[len(alltickers_df.index)] = [tick, cur_var, 0]
                    alltickers_df = alltickers_df.sort_values(by='Var')
                    alltickers_df.to_excel(alltickers_path, index=False)
                    coloring_portfolio(alltickers_path)
                    alltickers_df = pd.read_excel(alltickers_path)
                    alltickers_df.set_index("Symbol", drop=False, inplace=True)
                    portfolio_df.at[tick, 'Var'] = alltickers_df.at[tick, 'Var']
                    portfolio_df.at[tick, 'Qual'] = alltickers_df.at[tick, 'Qual']
                except Exception as e:
                    logger.error(f"Failed to calculate VaR for {tick}: {e}")
                    portfolio_df.at[tick, 'Var'] = 0
                    portfolio_df.at[tick, 'Qual'] = 'UNKNOWN'

        portfolio_df = portfolio_df.sort_values(by='Var')
        alltickers_df = alltickers_df.sort_values(by='Var')
        alltickers_df.to_excel(alltickers_path, index=False)
        portfolio_df.to_excel(portfolio_path, index=False)
        logger.info("VaR calculation completed")

    except Exception as e:
        logger.error(f"Error in get_var: {e}")
        raise


if __name__ == '__main__':
    try:
        logger.info("Starting portfolio processing")
        # Transform raw portfolio CSV to standardized format
        transform_df(config.PORTFOLIO_PATH_ORIGINAL, config.NET_LIQUIDITY)
        # Calculate VaR and add quality ratings
        get_var(config.PORTFOLIO_PATH_TRANSFORMED, config.ALL_TICKERS_PATH)
        logger.info("Portfolio processing completed successfully")
    except Exception as e:
        logger.error(f"Portfolio processing failed: {e}", exc_info=True)
        raise
